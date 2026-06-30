import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BANK_CONFIGS = [
    {
        "key": "full",
        "label": "全量题库",
        "input": Path(r"C:\Users\14274\Downloads\毛概系统题库_合并终版.xlsx"),
        "id_offset": 200000,
        "expected": 786,
    },
    {
        "key": "core",
        "label": "核心重点题库",
        "input": Path(r"C:\Users\14274\Downloads\毛概重点题库_章节题型整理版.xlsx"),
        "id_offset": 300000,
        "expected": 154,
    },
]

OUTPUT_JS = Path(__file__).with_name("data.js")
TEMPLATE_XLSX = Path(__file__).with_name("题库导入模板.xlsx")

DIFFICULTY_LAYERS = ["基础层", "进阶层", "冲刺层"]
QUESTION_TYPES = ["单选", "多选", "判断", "简答"]
OPTION_LETTERS = list("ABCDEF")


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def split_path(value):
    text = clean(value)
    if not text:
        return ["未分章节"]
    return [p.strip() for p in re.split(r"[/／>＞\\|]+", text) if p.strip()] or ["未分章节"]


def split_tags(value):
    text = clean(value)
    if not text:
        return []
    return [p.strip() for p in re.split(r"[、,，;/；|]+", text) if p.strip()]


def normalize_type(value):
    text = clean(value)
    if "简答" in text:
        return "简答"
    if "判断" in text:
        return "判断"
    if "多选" in text or "不定项" in text:
        return "多选"
    if "单选" in text or text:
        return "单选"
    return ""


def strip_option_prefix(value, letter):
    text = clean(value)
    if not text:
        return ""
    return re.sub(rf"^{letter}\s*[\.．、:：]\s*", "", text, flags=re.I)


def answer_letters(answer, qtype):
    text = clean(answer).upper()
    if qtype == "简答":
        return []
    if qtype == "判断":
        if text in {"A", "TRUE"} or "对" in text or "正确" in text:
            return ["A"]
        if text in {"B", "FALSE"} or "错" in text or "错误" in text:
            return ["B"]
    letters = re.findall(r"[A-F]", text)
    seen = []
    for letter in letters:
        if letter not in seen:
            seen.append(letter)
    return seen


def row_dict(ws, row):
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    return {headers[c - 1]: ws.cell(row, c).value for c in range(1, ws.max_column + 1)}


def find_question_sheet(wb):
    required = {"章节", "题型", "题干", "A", "B", "C", "D", "正确答案"}
    for ws in wb.worksheets:
        headers = {clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}
        if required.issubset(headers):
            return ws
    raise ValueError("未找到包含题库字段的工作表")


def validate_question(q, row_no, bank_label):
    errors = []
    if not q["categoryPath"]:
        errors.append(f"{bank_label} 第{row_no}行章节不能为空")
    if q["questionType"] not in QUESTION_TYPES:
        errors.append(f"{bank_label} 第{row_no}行题型填写错误")
    if not q["stem"]:
        errors.append(f"{bank_label} 第{row_no}行题干不能为空")
    if q["questionType"] in {"单选", "多选"}:
        for letter in "ABCD":
            if not q["options"].get(letter):
                errors.append(f"{bank_label} 第{row_no}行选项{letter}不能为空")
        if not q["answerLetters"]:
            errors.append(f"{bank_label} 第{row_no}行正确答案应填写选项字母")
    if q["questionType"] == "判断" and not q["answerLetters"]:
        errors.append(f"{bank_label} 第{row_no}行判断题正确答案应填写正确/错误或对/错")
    if q["questionType"] == "简答" and not q["answerText"]:
        errors.append(f"{bank_label} 第{row_no}行简答题正确答案不能为空")
    return errors


def normalize_row(raw, row_index, bank_key, bank_label, id_offset):
    qtype = normalize_type(raw.get("题型"))
    chapter = split_path(raw.get("章节"))
    answer = clean(raw.get("正确答案"))
    options = {}
    for i, letter in enumerate(OPTION_LETTERS):
        value = raw.get(letter) if letter in raw else raw.get(f"选项{letter}")
        value = strip_option_prefix(value, letter)
        if value:
            options[letter] = value
    if qtype == "判断":
        options = {
            "A": options.get("A") or "正确",
            "B": options.get("B") or "错误",
        }
    source_id = clean(raw.get("题号")) or f"{bank_key.upper()}{row_index:03d}"
    return {
        "id": id_offset + row_index,
        "sourceId": source_id,
        "bankScope": bank_key,
        "bankLabel": bank_label,
        "difficultyLayer": "基础层",
        "categoryPath": chapter,
        "categoryKey": " / ".join(chapter),
        "questionType": qtype,
        "stem": clean(raw.get("题干")),
        "options": options,
        "answerLetters": answer_letters(answer, qtype),
        "answerText": answer,
        "explanation": clean(raw.get("解析")),
        "tags": split_tags(raw.get("易错点标签") or raw.get("知识点标签")),
        "source": clean(raw.get("来源/依据")),
        "autoScore": qtype != "简答",
    }


def load_bank(config):
    wb = openpyxl.load_workbook(config["input"], data_only=True)
    ws = find_question_sheet(wb)
    questions = []
    errors = []
    for row in range(2, ws.max_row + 1):
        raw = row_dict(ws, row)
        if not any(clean(v) for v in raw.values()):
            continue
        q = normalize_row(raw, len(questions) + 1, config["key"], config["label"], config["id_offset"])
        row_errors = validate_question(q, row, config["label"])
        if row_errors:
            errors.extend(row_errors)
            continue
        questions.append(q)
    if errors:
        raise ValueError("\n".join(errors[:80]))
    if config.get("expected") and len(questions) != config["expected"]:
        raise ValueError(f"{config['label']} 题量应为 {config['expected']}，实际为 {len(questions)}")
    return questions


def count_by(items, field):
    counts = {}
    for q in items:
        counts[q[field]] = counts.get(q[field], 0) + 1
    return counts


def build_meta(banks):
    bank_scopes = [
        {"key": cfg["key"], "label": cfg["label"], "count": len(banks[cfg["key"]])}
        for cfg in BANK_CONFIGS
    ]
    meta = {
        "title": "毛泽东思想和中国特色社会主义理论体系概论刷题系统",
        "questionCount": sum(len(items) for items in banks.values()),
        "defaultBankScope": "full",
        "bankScopes": bank_scopes,
        "difficultyLayers": DIFFICULTY_LAYERS,
        "questionTypes": QUESTION_TYPES,
        "generatedFrom": {cfg["key"]: cfg["input"].name for cfg in BANK_CONFIGS},
        "banks": {},
    }
    for cfg in BANK_CONFIGS:
        key = cfg["key"]
        questions = banks[key]
        meta["banks"][key] = {
            "label": cfg["label"],
            "questionCount": len(questions),
            "layerCounts": {layer: 0 for layer in DIFFICULTY_LAYERS},
            "typeCounts": {qtype: 0 for qtype in QUESTION_TYPES},
            "chapterCounts": {},
        }
        meta["banks"][key]["layerCounts"]["基础层"] = len(questions)
        meta["banks"][key]["typeCounts"].update(count_by(questions, "questionType"))
        meta["banks"][key]["chapterCounts"] = count_by(
            [{"chapter": q["categoryKey"]} for q in questions],
            "chapter",
        )
    return meta


def write_data_js(banks):
    meta = build_meta(banks)
    chunks = [
        "// Auto-generated from the Excel question banks. Do not edit by hand.",
        f"window.QUESTION_BANKS = {json.dumps(banks, ensure_ascii=False)};",
        "window.QUESTION_BANK = window.QUESTION_BANKS.full;",
        f"window.QUESTION_META = {json.dumps(meta, ensure_ascii=False)};",
        "window.QUESTION_SCHEMA = {difficultyLayers:['基础层','进阶层','冲刺层'],questionTypes:['单选','多选','判断','简答']};",
        "",
    ]
    OUTPUT_JS.write_text("\n".join(chunks), encoding="utf-8")


def write_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题目导入模板"
    headers = ["章节", "题型", "题干", "A", "B", "C", "D", "正确答案", "解析", "易错点标签"]
    ws.append(headers)
    samples = [
        ["导论 马克思主义中国化时代化", "单选", "马克思主义中国化第一次历史性飞跃产生的理论成果是（ ）。", "毛泽东思想", "邓小平理论", "“三个代表”重要思想", "科学发展观", "A", "毛泽东思想是马克思主义中国化第一次历史性飞跃的理论成果。", "第一次飞跃"],
        ["第一章 毛泽东思想及其历史地位", "多选", "毛泽东思想活的灵魂包括（ ）。", "实事求是", "群众路线", "独立自主", "依法治国", "ABC", "实事求是、群众路线、独立自主是毛泽东思想活的灵魂。", "活的灵魂"],
        ["第二章 新民主主义革命理论", "判断", "新民主主义革命是无产阶级领导的人民大众的反帝反封建革命。", "正确", "错误", "", "", "正确", "新民主主义革命领导阶级是无产阶级。", "革命性质"],
    ]
    for row in samples:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="DCEBFF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    widths = [28, 12, 56, 22, 22, 22, 22, 14, 56, 24]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    guide = wb.create_sheet("填写规范")
    guide.append(["字段", "是否必填", "填写规范"])
    guide_rows = [
        ["章节", "是", "章节名称，可用 / 分隔多级分类，例如：第一章/第一节"],
        ["题型", "是", "只能填写：单选、多选、判断、简答"],
        ["题干", "是", "题目正文"],
        ["A-D", "单选/多选必填", "判断题可填写 正确/错误；简答题可留空"],
        ["正确答案", "是", "单选填 A/B/C/D；多选填 ABC/ABD；判断填 正确/错误 或 对/错；简答填参考答案文本"],
        ["解析", "否", "题目解析或背诵要点"],
        ["易错点标签", "否", "可用 、 或 , 分隔多个标签"],
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 18
    guide.column_dimensions["C"].width = 80
    wb.save(TEMPLATE_XLSX)


def main():
    banks = {cfg["key"]: load_bank(cfg) for cfg in BANK_CONFIGS}
    write_data_js(banks)
    write_template()
    for cfg in BANK_CONFIGS:
        print(f"{cfg['label']}: {len(banks[cfg['key']])} questions from {cfg['input'].name}")
    print(f"Generated {OUTPUT_JS}.")
    print(f"Generated {TEMPLATE_XLSX}.")


if __name__ == "__main__":
    main()
